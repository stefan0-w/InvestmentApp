import openpyxl
from datetime import datetime
from django.utils.timezone import make_aware
from ..models import Asset, Transaction


def parse_excel_date(cell_value):
    """
    Obsługuje dwa formaty:
    - tekst "2025-01-10 14:20:00"
    - excelowy numer daty (float)
    """

    if isinstance(cell_value, datetime):
        # Excel czasami zwraca datetime wprost
        return make_aware(cell_value)

    if isinstance(cell_value, (int, float)):
        # Excel stores date as number since 1899-12-30
        base_date = datetime(1899, 12, 30)
        parsed = base_date + timedelta(days=float(cell_value))
        return make_aware(parsed)

    if isinstance(cell_value, str):
        try:
            parsed = datetime.strptime(cell_value.strip(), "%Y-%m-%d %H:%M:%S")
            return make_aware(parsed)
        except ValueError:
            raise ValueError(f"Nieprawidłowy format daty: {cell_value}")

    raise ValueError(f"Nieobsługiwany typ daty: {cell_value}")
    


def import_xtb_xlsx(file, user):
    """
    Import transakcji z uproszczonego pliku XLSX.
    """
    wb = openpyxl.load_workbook(file)
    ws = wb.active  # arkusz Transactions

    header = [cell.value for cell in ws[1]]

    required_cols = ["Symbol", "Name", "Quantity", "Price", "Type", "Date"]
    for col in required_cols:
        if col not in header:
            raise ValueError(f"Brak wymaganej kolumny: {col}")

    col_index = {name: header.index(name) for name in required_cols}

    portfolio = user.portfolio

    for row in ws.iter_rows(min_row=2):
        symbol = row[col_index["Symbol"]].value
        name = row[col_index["Name"]].value
        quantity = row[col_index["Quantity"]].value
        price = row[col_index["Price"]].value
        ttype = row[col_index["Type"]].value
        date_raw = row[col_index["Date"]].value

        if not symbol:
            continue

        transaction_date = parse_excel_date(date_raw)

        # asset
        asset, _ = Asset.objects.get_or_create(
            symbol=symbol,
            defaults={"name": name, "type": "STOCK"}
        )

        # rekord transakcji
        Transaction.objects.create(
            portfolio=portfolio,
            asset=asset,
            quantity=quantity,
            price=price,
            transaction_type=ttype.upper(),
            transaction_date=transaction_date,
            broker="XTB",
            source="import_xlsx",
        )

    return True
